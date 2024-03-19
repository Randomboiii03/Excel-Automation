import pandas as pd
from openpyxl.styles import PatternFill, Font
from openpyxl import load_workbook
from fuzzywuzzy import fuzz




def highlight_n_check_prediction(excel_file_path):

    def compare_address(str1, address, threshold=60):
        similarity_ratio = fuzz.token_set_ratio(str(str1).lower(), str(address).lower())

        new_ratio = similarity_ratio <= threshold
        if new_ratio:
            check = str1.replace(' ', '').replace('Ñ', 'N').lower() not in address.replace(' ', '').replace('Ñ', 'N').lower()
            if not check:
                new_ratio = check

        return new_ratio

    df = pd.read_excel(excel_file_path)

    # Load the workbook
    book = load_workbook(excel_file_path)

    # Access the active sheet
    sheet = book.active

    for row_index, row in df.iterrows():
        column_index1 = df.columns.get_loc('AREA') + 1
        cell1= sheet.cell(row=row_index + 2, column=column_index1)

        column_index2 = df.columns.get_loc('MUNICIPALITY') + 1
        cell2= sheet.cell(row=row_index + 2, column=column_index2)
        
        if compare_address(row["AREA"], row["ADDRESS"]):
            cell1.fill = PatternFill(start_color="fffa00", end_color="fffa00", fill_type="solid")

        if compare_address(row["MUNICIPALITY"], row["ADDRESS"]):
            cell2.fill = PatternFill(start_color="fffa00", end_color="fffa00", fill_type="solid")

        if compare_address(row["MUNICIPALITY"], row["ADDRESS"]) and compare_address(row["AREA"], row["ADDRESS"]):
            cell1.fill = PatternFill(start_color="ffa500", end_color="ffa500", fill_type="solid")
            cell2.fill = PatternFill(start_color="ffa500", end_color="ffa500", fill_type="solid")

    book.save('uploads\\new.xlsx')

if __name__ == "__main__":
    highlight_n_check_prediction('uploads\\Output-BPI-2024-03-19-0679.xlsx')