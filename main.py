import os
import pandas as pd
from fuzzywuzzy import fuzz
from flask import Flask, render_template, request
import random
from datetime import datetime
import json
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

app = Flask(__name__)

def get_template_header(file_path):
    return pd.read_excel(file_path).columns.to_list()

def compare_string(str1, str2, threshold=90):
    similarity_ratio = fuzz.token_set_ratio(str(str1).lower(), str(str2).lower())
    return similarity_ratio >= threshold

def find_header(row_values, template_header):
    row_list = row_values.tolist()
    for header in template_header:
        for data in row_list:
            if compare_string(data, header):
                return True
    return False

def get_index_of_header(excel_file_path, template_header) -> tuple[int, list]:
    work_book = pd.read_excel(excel_file_path, sheet_name=None, header=None)
    for sheet_name, sheet_data in work_book.items():
        if sheet_data.empty:
            return 0, []
        for index, row in sheet_data.iterrows():
            if find_header(row.values, template_header):
                return index, row.values.tolist()
    return 0, [""] * len(template_header)

def map_header(header, mapping):
    for key, value in mapping.items():
        if compare_string(header.lower(), key.lower()):
            return value
    return header
# #GETTING ONLY FIRST 6 CHAR IN CH CODE
# # def fill_missing_values(excel_file_path, campaign_file_path):
# #     with open(campaign_file_path, 'r') as file:
# #         campaign_data = json.load(file)
    
# #     excel_data = pd.read_excel(excel_file_path)
# #     for index, row in excel_data.iterrows():
# #         if pd.isna(row['BANK']):
# #             ch_code = str(row['CH CODE'])[:6]  # Get first 6 characters of CH CODE
# #             if ch_code in campaign_data:
# #                 excel_data.at[index, 'BANK'] = campaign_data[ch_code]['BANK']
# #         if pd.isna(row['PLACEMENT']):
# #             ch_code = str(row['CH CODE'])[:6]  # Get first 6 characters of CH CODE
# #             if ch_code in campaign_data:
# #                 excel_data.at[index, 'PLACEMENT'] = campaign_data[ch_code]['PLACEMENT']
    
# #     excel_data.to_excel(excel_file_path, index=False)


# # GETTING ONLY FIRST NUMBER AND LETTERS IN CH CODE
# def fill_missing_values(excel_file_path, campaign_file_path):
#     with open(campaign_file_path, 'r') as file:
#         campaign_data = json.load(file)
    
#     excel_data = pd.read_excel(excel_file_path)
#     for index, row in excel_data.iterrows():
#         if pd.isna(row['BANK']):
#             ch_code_match = re.search(r'^(\d+[A-Z]+)', str(row['CH CODE']))  # Extract first numbers and letters from CH CODE
#             if ch_code_match is not None:
#                 ch_code = ch_code_match.group(1)
#                 if ch_code in campaign_data:
#                     excel_data.at[index, 'BANK'] = campaign_data[ch_code]['BANK']
#         if pd.isna(row['PLACEMENT']):
#             ch_code_match = re.search(r'^(\d+[A-Z]+)', str(row['CH CODE']))  # Extract first numbers and letters from CH CODE
#             if ch_code_match is not None:
#                 ch_code = ch_code_match.group(1)
#                 if ch_code in campaign_data:
#                     excel_data.at[index, 'PLACEMENT'] = campaign_data[ch_code]['PLACEMENT']
    
#     excel_data.to_excel(excel_file_path, index=False)

 # GETTING ONLY FIRST NUMBER AND LETTERS IN CH CODE || EXCLUDE ALL STRING VALUE
def fill_missing_values(excel_file_path, campaign_file_path):
    with open(campaign_file_path, 'r') as file:
        campaign_data = json.load(file)
    
    excel_data = pd.read_excel(excel_file_path)
    # Convert CH CODE to string
    excel_data['CH CODE'] = excel_data['CH CODE'].astype(str)
    # Drop rows with CH CODE that has all-letter value
    excel_data = excel_data[~excel_data['CH CODE'].str.isalpha()]
    
    for index, row in excel_data.iterrows():
        if pd.isna(row['BANK']):
            ch_code_match = re.search(r'^(\d+[A-Z]+)', str(row['CH CODE']))  # Extract first numbers and letters from CH CODE
            if ch_code_match is not None:
                ch_code = ch_code_match.group(1)
                if ch_code in campaign_data:
                    excel_data.at[index, 'BANK'] = campaign_data[ch_code]['BANK']
        if pd.isna(row['PLACEMENT']):
            ch_code_match = re.search(r'^(\d+[A-Z]+)', str(row['CH CODE']))  # Extract first numbers and letters from CH CODE
            if ch_code_match is not None:
                ch_code = ch_code_match.group(1)
                if ch_code in campaign_data:
                    excel_data.at[index, 'PLACEMENT'] = campaign_data[ch_code]['PLACEMENT']
    
    excel_data.to_excel(excel_file_path, index=False)

def extract_address(output_file_path, address_column_name, output_address_file_path):
    output_data = pd.read_excel(output_file_path)
    address_data = output_data[address_column_name].dropna()
    address_data.to_excel(output_address_file_path, index=False)


def auto_fit_columns(excel_file_path):
    wb = load_workbook(excel_file_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for column_cells in ws.columns:
            # Apply style to the header cells
            header_fill = PatternFill(start_color="D4AC0D", end_color="D4AC0D", fill_type="solid")  # Gray color
            header_font = Font(bold=True)
            for cell in column_cells:
                if cell.row == 1:  # Assuming header is in the first row
                    cell.fill = header_fill
                    cell.font = header_font
                    # Convert header text to uppercase
                    cell.value = str(cell.value).upper()
            # Auto-fit column width
            max_length = max(len(str(cell.value)) for cell in column_cells)
            adjusted_width = (max_length + 2) * 1.2  # Adjusted width with a factor for padding
            ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width
    wb.save(excel_file_path)





@app.route('/', methods=['GET', 'POST'])
def index():
    message = None
    bank_names = []
    for dirpath, dirnames, filenames in os.walk("C:\\Users\\SPM\Desktop\\ONLY SAVE FILE HERE\\SIT-OJT\\Requests"):
        bank_names.extend(dirnames)
        break  # Only the top-level directories are needed

    if request.method == 'POST':
        bank_name = request.form['bank_name']
        folder_path = os.path.join("C:\\Users\\SPM\Desktop\\ONLY SAVE FILE HERE\\SIT-OJT\\Requests", bank_name)
        files = [f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))]
        template_header = get_template_header("Template.xlsx")
        datas = [template_header]

        mapping = {
            "REQUEST DATE": "DATE REQUESTED",
            "REQUEST NAME": "REQUESTED BY"
            # Add more mappings as needed
        }

        existing_headers = set()

        for file in files:
            excel_file_path = os.path.join(folder_path, file)
            index_header, current_header = get_index_of_header(excel_file_path, template_header)
            work_book = pd.read_excel(excel_file_path, sheet_name=None, header=index_header)
            for sheet_name, sheet_data in work_book.items():
                for index, row in sheet_data.iterrows():
                    output_row = []
                    for header in template_header:
                        value = None
                        for col_header, col_value in row.items():
                            if compare_string(header, col_header):
                                value = col_value
                                break
                        output_row.append(value)
                    datas.append(output_row)

                   # Append missing columns from the file
                    for col_header in row.keys():
                        mapped_header = map_header(col_header, mapping)
                        # if mapped_header.lower() not in [h.lower() for h in template_header] and mapped_header not in existing_headers and compare_string(mapped_header, col_header):
                        #     datas[0].append(mapped_header)
                        #     output_row.append(row[col_header])
                        #     existing_headers.add(mapped_header)
                        if mapped_header.lower() not in [h.lower() for h in template_header] and mapped_header not in existing_headers and not (col_header.strip() == "" or col_header.startswith("Unnamed")) and compare_string(mapped_header, col_header):
                            datas[0].append(mapped_header)
                            output_row.append(row[col_header])
                            existing_headers.add(mapped_header)

        output_work_book = pd.DataFrame(datas[1:], columns=datas[0])
        random_number = "".join([str(random.randint(0, 9)) for _ in range(4)])
        current_date = datetime.now().strftime("%Y-%m-%d")
        output_file_name = f"Output-{bank_name}-{current_date}-{random_number}.xlsx"
        output_file_path = os.path.join("C:\\Users\\SPM\Desktop\\ONLY SAVE FILE HERE\\SIT-OJT\\Merged", output_file_name)
        output_work_book.to_excel(output_file_path, index=False)

        campaign_file_path = 'campaign_list.json' 
        fill_missing_values(output_file_path, campaign_file_path)

        address_column_name = "ADDRESS"  # Replace with the actual column name containing the address
        output_address_file_name = f"Output-Address-{bank_name}-{current_date}-{random_number}.xlsx"
        output_address_file_path = os.path.join("C:\\Users\\SPM\Desktop\\ONLY SAVE FILE HERE\\SIT-OJT\\Merged\\Area", output_address_file_name)
        extract_address(output_file_path, address_column_name, output_address_file_path)

        message = f"Excel file created successfully for {bank_name}. Output file: <strong>{output_file_name}</strong>. Address file: <strong>{output_address_file_name}</strong>"

          # Call auto_fit_columns function after saving the Excel file
        auto_fit_columns(output_file_path)
        auto_fit_columns(output_address_file_path)
        

    return render_template('index.html', message=message, bank_names=bank_names)

if __name__ == '__main__':
    app.run(debug=True)



