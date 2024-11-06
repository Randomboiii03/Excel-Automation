from fuzzywuzzy import fuzz
import json
import re
from openpyxl.styles import PatternFill, Font
from openpyxl import load_workbook
import pandas as pd
import polars as pl
import streamlit as st

def get_index_of_header(excel_file_path, template_header) -> int:
    sheet_data = pd.read_excel(excel_file_path, sheet_name=0, header=None)
    if sheet_data.empty:
        return -1
    for index, row in sheet_data.iterrows():
        if find_header(row.values, template_header):
            return index
    return -1
    

def find_header(row_values, template_header):
    for data in row_values:
        for header in template_header:
            if compare_string(data, header):
                return True
    return False


def clean_string(input_string):
    return re.sub(r'[^a-zA-Z0-9\s]', ' ', input_string).upper()


def compare_string(str1, str2, threshold=90):
    similarity_ratio = fuzz.token_set_ratio(clean_string(str(str1)).lower(), clean_string(str(str2)).lower())
    return similarity_ratio >= threshold


def map_header(mapping, header):
    cleaned_header = clean_string(header)
    return mapping.get(cleaned_header, None)


def is_empty_or_spaces(s):
    return s.replace(' ', '') == ''


def fill_bank_n_chcode(df, campaign_file_path):
    with open(campaign_file_path, 'r') as file:
        campaign_data = json.load(file)

    placeholder3 = st.empty()
    placeholder3.text(f"Checking Ch Code 🔃")
    
    try:
        ch_codes = pl.Series(df.with_columns(pl.col('CH CODE').str.len_chars().alias("n_chars")
                                            ).filter((pl.col('CH CODE').is_not_null()) & 
                                                    (pl.col('n_chars') > 3) &
                                                    ((pl.col('BANK').is_null()) |
                                                    (pl.col('PLACEMENT').is_null()))
                                                    ).drop(["n_chars"]).select(pl.col("CH CODE"))).unique().to_list()
    except:
        placeholder3.text(f"Failed to fill Bank and Placement ❌")
        return df
    
    if not ch_codes:
        placeholder3.text(f"No Bank and Placement to fill❌")
        return df

    banks, placements = zip(*[
        (campaign_data.get(ch_code_match.group(1), {'BANK': '', 'PLACEMENT': ''})['BANK'],
        campaign_data.get(ch_code_match.group(1), {'BANK': '', 'PLACEMENT': ''})['PLACEMENT'])
        for ch_code in ch_codes
        if (ch_code_match := re.search(r'^(\d+[A-Z]+)', ch_code))
    ]) if ch_codes else ([], [])


    ch_codes_df = pl.DataFrame([banks, placements], schema=["BANK", "PLACEMENT"]).insert_column(0, pl.Series("CH CODE", ch_codes))

    placeholder3.text(f"Bank and Placement Filled Up ✅")

    return df.join(ch_codes_df, on='CH CODE', how='left'
                 ).with_columns([
                     pl.col('BANK').fill_null(pl.col('BANK_right')), 
                     pl.col('PLACEMENT').fill_null(pl.col('PLACEMENT_right'))
                     ]).drop(['BANK_right', 'PLACEMENT_right'])


def compare_address(str1, address, threshold=60):
        if str1.replace(' ', '').replace('ñ', 'n').lower() in address.replace(' ', '').replace('ñ', 'n').lower():
            return False

        similarity_ratio = fuzz.token_set_ratio(clean_string(str(str1)).lower(), clean_string(str(address)).lower())
        return similarity_ratio <= threshold


def highlighting_excel(excel_file_path, type="merge"):
    placeholder1 = st.empty()
    placeholder2 = st.empty()
    placeholder3 = st.empty()

    if type == "merge":
        placeholder1.text(f"Highlighting blank Ch Code and Campaign 🔃")
    placeholder2.text(f"Highlighting predictions 🔃")

    df = pd.read_excel(excel_file_path)
    book = load_workbook(excel_file_path)
    sheet = book.active

    if type == "merge":
        col_list = ['BANK', 'PLACEMENT', 'AREA', 'MUNICIPALITY']
    else:
        col_list = ['AREA', 'MUNICIPALITY']

    column_indices = {col: df.columns.get_loc(col) + 1 for col in col_list}
    
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    red_fill = PatternFill(start_color='FF4400', end_color='FF4400', fill_type='solid')
    light_yellow_fill = PatternFill(start_color='FFFA00', end_color='FFFA00', fill_type='solid')
    header_fill = PatternFill(start_color='D4AC0D', end_color='D4AC0D', fill_type='solid')
    header_font = Font(bold=True)

    for row_index, row in df.iterrows():
        row_index += 2

        if type == "merge":
            bank = str(row.get('BANK', '')).lower()
            placement = str(row.get('PLACEMENT', '')).lower()

            cell1 = sheet.cell(row=row_index, column=column_indices['BANK'])
            cell2 = sheet.cell(row=row_index, column=column_indices['PLACEMENT'])

            if bank == "nan":
                cell1.fill = yellow_fill
            elif placement == "nan":
                cell2.fill = yellow_fill

        address = str(row.get('ADDRESS', '')).lower()
        area = str(row.get('AREA', '')).lower()
        municipality = str(row.get('MUNICIPALITY', '')).lower()

        cell3 = sheet.cell(row=row_index, column=column_indices['AREA'])
        cell4 = sheet.cell(row=row_index, column=column_indices['MUNICIPALITY'])
        
        if not address.strip() or (area == "nan" and municipality == "nan"):
            cell3.fill = cell4.fill = red_fill
        else:
            if area == "ncr":
                area = "manila"

            area_match = compare_address(area, address)
            municipality_match = compare_address(municipality, address)

            if area_match and municipality_match:
                cell3.fill = cell4.fill = orange_fill
            elif area_match:
                cell3.fill = light_yellow_fill
            elif municipality_match:
                cell4.fill = light_yellow_fill

    if type == "merge":
        placeholder1.text(f"Blank Ch Code and Campaign Highlighted ✅")
    placeholder2.text(f"Predictions highlighted ✅")
    placeholder3.text(f"Highlighting headers 🔃")

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.value = str(cell.value).upper()

    placeholder3.text(f"Headers highlighted ✅")

    return book
